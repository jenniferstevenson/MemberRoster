#!/usr/bin/env python3

from robobrowser import RoboBrowser
import zipfile as zf
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
import codecs
import sys
from PyQt5 import QtWidgets
import UserInterface
from datetime import date

"""

EDIT THE FOLLOWING LISTS OF SPG AND LIDN IDs

"""

SPG_IDs = ["CA2043", "FL0343", "FL2221", "MD0094", "658406", "CA2700", "MN2013", "PA2205", "IA2053", "MI2035", "649419", "MI2002", "OH2233", "TX0155", "KY5005", "WA0014", "IL2464", "MO0014", "TX2246", "635796", "CA0053", "616890", "CA053", "601045", "OH2004", "669907", "700273", "631225", "IL2185"]

LIDN_IDs = ["CA2043", "658406", "MN2013", "MI2035", "649419", "TX0155", "KY5005", "IL5043", "635796", "CA0053", "616890", "CA053", "601045"]


"""
DO NOT EDIT THE REMAINDER OF THE FILE

"""

url = "https://legacy.premierinc.com/bp/hipaa?from=%2Fabout%2Fprivate%2Fsuppliers%2Fcontracted-suppliers%2Frosters%2Fdisplay.jsp&roster_type=hisci&cn=master_hin"


def main():
    app = QtWidgets.QApplication([])
    window = UserInterface.UI()
    window.show()

    for fileName in os.listdir("."):
        if fileName.endswith(".xlsx"):
            # delete unused files
            os.remove(fileName)

    window.show_welcome_message()
    sys.exit(app.exec_())


def run_script(username, password, ui):
    # open the Premier Connect applet
    browser = RoboBrowser(parser="html.parser")
    browser.open(url)

    # login to the page
    login(browser, "login-form", username, password)

    # follow the link to the latest files
    link = find_link(browser, ui)

    if not link:
        return
    browser.follow_link(link)

    # download and extract the zip file
    download_zip(browser.response)

    extract_zip()

    today = date.today()
    excel_filename = "MemberRoster_" + str(today.month) + "_" + str(today.day) + ".xlsx"

    # iterate over files, looking for combined .txt file
    for fileName in os.listdir("."):
        if fileName.startswith("Premier_HISCI_Roster_W_HIN_Combined_"):
            # create a new workbook
            wb = Workbook()

            # create the 4 sheets
            create_sheets(wb, fileName)

            # save the file
            wb.save(excel_filename)

            # delete remaining files
            os.remove(fileName)
            os.remove("output.exe")

        elif fileName.endswith(".txt"):
            # delete unused files
            os.remove(fileName)

    # open the excel file
    if os.name == "nt":
        os.system(excel_filename)
    elif os.name == "posix":
        os.system("open " + excel_filename)


def find_link(browser, ui):
    # get all roster links and return the most recent
    links = browser.get_links(text="Premier_HISCI_Roster_W_HIN_")

    if len(links) == 0:
        ui.login_failed()
        return False
    else:
        return links[0]


def login(browser, form_name, username, password):
    # get the login form, set the username and password and login
    login_form = browser.get_form(class_=form_name)
    login_form["username"].value = username
    login_form["password"].value = password
    login_form.serialize()
    browser.submit_form(login_form)


def download_zip(response):
    # write each line of the binary to output.exe
    with open("output.exe", "wb") as f:
        f.write(response.content)


def extract_zip():
    # extract all files in output.exe
    with zf.ZipFile("output.exe", "r") as myFile:
        myFile.extractall()


def create_sheets(wb, file_name):
    # iterate over ID numbers, creating a new sheet for each
    ws = wb.active
    ws.title = "SPG OLM"
    worksheets = [ws, wb.create_sheet("SPG Aff"), wb.create_sheet("LIDN OLM"), wb.create_sheet("LIDN Aff")]

    first_row_appended = False

    with codecs.open(file_name, "r", encoding="utf-8", errors="ignore") as f:
        # iterate through .txt file
        for line in f:
            # remove quote characters
            line = line.replace('"', "")

            # remove new line character
            line = line.replace('\n', "")

            # split by tabs
            row = line.split("\t")

            # choose only active rows
            status = row[21]
            if status != "Active" and first_row_appended:
                continue

            # remove unnecessary columns
            row = row[:26]
            remove_columns([1, 2, 13, 14, 15, 17, 18, 21, 22, 23, 24], row)

            if first_row_appended:
                # append row to correct sheet
                append_rows(worksheets, row)
            elif row[0] == "GPO ID":
                row.insert(-1, "Top Parent Affiliate")
                # append first row to all sheets
                worksheets[0].append(row)
                worksheets[1].append(row)
                worksheets[2].append(row)
                worksheets[3].append(row)
                first_row_appended = True


    # style the columns
    style_columns(worksheets)


def append_rows(worksheets, row):
    TP_ID = row[12]
    RELATIONSHIP = row[11]

    # add top parent name to affiliates
    if TP_ID in ["635796", "CA0053", "616890", "CA053", "601045"]:
        row.insert(-1, "Adventist Health Aff")
    elif TP_ID in ["OH2004", "669907"]:
        row.insert(-1, "Mercy Health Aff")
    elif TP_ID == "700273":
        row.insert(-1, "Peace Health Aff")
    elif TP_ID in ["631225", "IL2185"]:
        row.insert(-1, "Presence Health Network Aff")
    else:
        row.insert(-1, "")

    # append row to correct sheet
    for spg_id in SPG_IDs:
        if spg_id == TP_ID and ("Owned" == RELATIONSHIP or "Leased" == RELATIONSHIP or "Managed" == RELATIONSHIP):
            worksheets[0].append(row)
        if spg_id == TP_ID and ("Affiliated" == RELATIONSHIP or "Employed" == RELATIONSHIP):
            worksheets[1].append(row)
    for lidn_id in LIDN_IDs:
        if lidn_id == TP_ID and ("Owned" == RELATIONSHIP or "Leased" == RELATIONSHIP or "Managed" == RELATIONSHIP):
            worksheets[2].append(row)
        if lidn_id == TP_ID and ("Affiliated" == RELATIONSHIP or "Employed" == RELATIONSHIP):
            worksheets[3].append(row)


def remove_columns(numbers, row):
    i = 0
    for num in numbers:
        row.pop(num - i)
        i += 1


def style_columns(worksheets):
    label_font = Font(bold=True)
    label_fill = PatternFill(fill_type="solid", fgColor="64A70B")
    label_alignment = Alignment(horizontal="center", vertical="center")

    for ws in worksheets:
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 36
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 25
        ws.column_dimensions["G"].width = 15
        ws.column_dimensions["H"].width = 15
        ws.column_dimensions["I"].width = 12
        ws.column_dimensions["J"].width = 12
        ws.column_dimensions["L"].width = 25
        ws.column_dimensions["M"].width = 16
        ws.column_dimensions["N"].width = 32
        ws.column_dimensions["O"].width = 25
        ws.column_dimensions["P"].width = 25
        ws.row_dimensions[1].height = 25

        for col in ws.iter_cols(min_row=0, max_row=0):
            col[0].font = label_font
            col[0].fill = label_fill
            col[0].alignment = label_alignment


if __name__ == "__main__":
    main()
